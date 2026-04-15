"""
TD 投放数据查询工具
通过 Ad Manager 报表 API 查询投放消耗、ROI、LTV、留存等数据

用法:
  python tools/td_query.py --report --body '{"page":1,...}' --format markdown
  python tools/td_query.py --hourly --body '{"date":"2026-03-02",...}' --format markdown
  python tools/td_query.py --download --report-type channel --body '{"dimension_list":[5,3],...}'
  python tools/td_query.py --download --report-type channelDetail --body '{"dimension_items":[3,16],...}'
"""

import argparse
import io
import json
import os
import sys
import tempfile
import time

import requests

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from utils import (load_user_config, format_json, format_markdown_table, check_td_token_expiry,
                   get_injected_credential, save_user_credentials, clear_user_credentials,
                   sync_credential_env, clear_credential_env)
from defaults import TD_CONFIG

ASYNC_TASK_CODE = 102241

HOURLY_TOTAL_LABELS = {
    "new_user_total": "新增设备",
    "pay_user_total": "付费用户",
    "pay1_total": "首日付费",
    "account_total": "账户数",
    "cost_total": "消耗",
    "show_total": "展示",
    "click_total": "点击",
    "active_total": "激活",
    "cost_per_new_user_total": "新增成本",
}


def get_token(config):
    """从环境变量或 config 获取 JWT token"""
    token = get_injected_credential("td", "token")
    if not token:
        token = os.environ.get("AD_MANAGER_TOKEN", "")
    if not token:
        td = config.get("td", {})
        token = td.get("token", "")
    if not token:
        print("Error: TD token not found. "
              "请告知用户发送「设置TD凭证」来配置 token。",
              file=sys.stderr)
        sys.exit(1)

    days, expire_time = check_td_token_expiry(token)
    if days is not None:
        if days <= 0:
            print(f"Warning: TD token has expired (at {expire_time}), "
                  f"请告知用户发送「设置我的TD账号」来更新 token。",
                  file=sys.stderr)
        elif days <= 1:
            hours = days * 24
            print(f"Warning: TD token expires in {hours:.0f}h "
                  f"(at {expire_time}), 请提醒用户发送「设置TD凭证」更新 token。",
                  file=sys.stderr)

    return token


def _post(url, token, body):
    """发送 POST 请求并返回解析后的 JSON"""
    resp = requests.post(
        url,
        json=body,
        headers={
            "authorization": token,
            "Content-Type": "application/json",
        },
        timeout=60,
    )
    if resp.status_code == 401:
        print("Error: Token expired or invalid (HTTP 401)", file=sys.stderr)
        sys.exit(1)
    resp.raise_for_status()
    data = resp.json()
    if data.get("code") != 0:
        code = data.get("code", "?")
        detail = data.get("detail", data.get("message", "unknown"))
        print(f"API error [code={code}]: {detail}", file=sys.stderr)
        print(format_json(data))
        sys.exit(1)
    return data


def _pick_display_columns(row, column_list):
    """从返回行中挑选维度字段 + 请求的指标字段作为显示列"""
    dim_cols = [
        f for f in TD_CONFIG["dimension_fields"]
        if row.get(f) not in (None, "", 0, "0")
    ]
    return dim_cols + (column_list or [])


def _format_value(val):
    """格式化数值：百分比、千分位、小数"""
    if val is None:
        return "-"
    if isinstance(val, bool):
        return str(val)
    if isinstance(val, int):
        return f"{val:,}"
    if isinstance(val, float):
        if abs(val) < 1:
            return f"{val * 100:.2f}%"
        return f"{val:,.2f}"
    return str(val) if val != "" else "-"


def _print_markdown_table(rows, columns):
    """输出 markdown 表格，数值经过格式化"""
    if not rows:
        print("(无数据)")
        return
    headers = columns
    formatted_rows = []
    for row in rows:
        formatted_rows.append([_format_value(row.get(c)) for c in columns])
    print(format_markdown_table(headers, formatted_rows))


def query_report(config, body_str, fmt, max_length):
    """常规报表查询"""
    token = get_token(config)
    body = json.loads(body_str)
    url = TD_CONFIG["report_url"]

    print(f"Querying TD report...", file=sys.stderr)
    result = _post(url, token, body)

    data = result["data"]
    rows = data.get("data_list", [])
    total = data.get("total", len(rows))
    summary = data.get("summary")

    if not rows:
        print("(无数据)")
        return

    columns = _pick_display_columns(rows[0], body.get("column_list"))

    if fmt == "json":
        output = format_json(result)
        if max_length > 0:
            output = output[:max_length] + "\n...(truncated)"
        print(output)
        return

    page = data.get("page", 1)
    print(f"\n共 {total} 条，当前第 {page} 页，展示 {len(rows)} 条:\n")
    _print_markdown_table(rows, columns)

    if summary:
        print("\n**汇总:**\n")
        _print_markdown_table([summary], columns)


def query_hourly(config, body_str, fmt, max_length):
    """分小时报表查询"""
    token = get_token(config)
    body = json.loads(body_str)
    url = TD_CONFIG["hourly_url"]

    print(f"Querying TD hourly report...", file=sys.stderr)
    result = _post(url, token, body)

    data = result["data"]
    rows = data.get("data_list", [])

    if not rows:
        print("(无数据)")
        return

    if fmt == "json":
        output = format_json(result)
        if max_length > 0:
            output = output[:max_length] + "\n...(truncated)"
        print(output)
        return

    # 自动检测列，date 和 hour 优先
    col_set = {}
    for row in rows:
        for k in row:
            col_set[k] = None
    priority = ["date", "hour"]
    columns = [p for p in priority if p in col_set]
    columns += [k for k in col_set if k not in priority]

    print(f"\n共 {len(rows)} 个小时数据:\n")
    _print_markdown_table(rows, columns)

    # 汇总信息
    lines = []
    for key, label in HOURLY_TOTAL_LABELS.items():
        val = data.get(key)
        if val is not None:
            lines.append(f"- **{label}**: {_format_value(val)}")
    if data.get("latest_time"):
        lines.append(f"- **数据更新时间**: {data['latest_time']}")
    if lines:
        print("\n**汇总:**\n")
        print("\n".join(lines))


def _post_download(url, token, body):
    """发送下载请求，返回 (content_type, data)"""
    resp = requests.post(
        url,
        json=body,
        headers={
            "authorization": token,
            "Content-Type": "application/json",
        },
        timeout=120,
    )
    if resp.status_code == 401:
        print("Error: Token expired or invalid (HTTP 401)", file=sys.stderr)
        sys.exit(1)
    resp.raise_for_status()
    ct = resp.headers.get("Content-Type", "")
    if "application/json" in ct:
        return "json", resp.json()
    return "binary", resp.content


def _poll_async_task(token, task_id, interval=3, max_wait=300):
    """轮询异步任务，返回下载 URL"""
    url = f"{TD_CONFIG['async_task_url']}{task_id}/"
    elapsed = 0
    fail_count = 0
    max_fail = 5
    print(f"Async task created (ID: {task_id}), polling...", file=sys.stderr)
    while elapsed < max_wait:
        time.sleep(interval)
        elapsed += interval
        resp = requests.get(
            url,
            headers={"authorization": token},
            timeout=60,
        )
        result = resp.json()
        code = result.get("code", -1)
        if code == 0:
            dl_url = result.get("data", {}).get("url", "")
            if dl_url:
                print(f"  [{elapsed}s] Task completed", file=sys.stderr)
                return dl_url
            print(f"  [{elapsed}s] Task completed but no download URL",
                  file=sys.stderr)
            sys.exit(1)
        elif code == 102244:
            fail_count += 1
            if fail_count >= max_fail:
                print(f"  [{elapsed}s] Task failed after {fail_count} "
                      f"consecutive errors: {result.get('detail', '')}",
                      file=sys.stderr)
                sys.exit(1)
            print(f"  [{elapsed}s] Task not ready (code=102244, "
                  f"retry {fail_count}/{max_fail})", file=sys.stderr)
        else:
            fail_count = 0
            print(f"  [{elapsed}s] Waiting... (code={code})", file=sys.stderr)
    print(f"Polling timeout ({max_wait}s)", file=sys.stderr)
    sys.exit(1)


def _download_from_url(url):
    """从 URL 下载文件内容"""
    resp = requests.get(url, timeout=300)
    resp.raise_for_status()
    return resp.content


LARGE_FILE_WARN_BYTES = 5 * 1024 * 1024  # 5 MB


def _parse_excel_to_rows(excel_bytes, max_rows=50, grep=None):
    """解析 Excel 字节流，返回 (headers, data_rows, total_rows)。

    grep: 关键词过滤，匹配任意列包含该关键词的行。
          启用时用流式解析扫描全部行，返回匹配行（上限 max_rows）。
    """
    size_mb = len(excel_bytes) / (1024 * 1024)
    if len(excel_bytes) > LARGE_FILE_WARN_BYTES:
        print(f"Warning: Large file ({size_mb:.1f} MB), "
              f"parsing may take a while...", file=sys.stderr)

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    total_rows = ws.max_row - 1
    headers = [cell.value or "" for cell in ws[1]]

    data_rows = []
    if grep:
        grep_lower = grep.lower()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(grep_lower in str(v or "").lower() for v in row):
                data_rows.append(dict(zip(headers, row)))
                if len(data_rows) >= max_rows:
                    break
        print(f"Grep \"{grep}\": {len(data_rows)} matches "
              f"in {total_rows} rows", file=sys.stderr)
    else:
        for row in ws.iter_rows(min_row=2,
                                max_row=min(max_rows + 1, ws.max_row),
                                values_only=True):
            data_rows.append(dict(zip(headers, row)))

    wb.close()

    return headers, data_rows, total_rows


def query_download(config, body_str, report_type, fmt, max_rows, grep=None):
    """下载报表查询（渠道聚合 / 渠道明细）"""
    token = get_token(config)
    body = json.loads(body_str)

    if report_type == "channel":
        url = TD_CONFIG["channel_download_url"]
    else:
        url = TD_CONFIG["detail_download_url"]

    print(f"Downloading TD {report_type} report...", file=sys.stderr)

    ct, data = _post_download(url, token, body)

    if ct == "json":
        code = data.get("code", -1)
        if code == ASYNC_TASK_CODE:
            task_id = data.get("data", {}).get("id")
            if not task_id:
                print("Error: Async task created but no task ID",
                      file=sys.stderr)
                sys.exit(1)
            dl_url = _poll_async_task(token, task_id)
            excel_bytes = _download_from_url(dl_url)
        elif code == 0:
            print("(无数据)")
            return
        else:
            detail = data.get("detail", data.get("message", "unknown"))
            print(f"API error [code={code}]: {detail}", file=sys.stderr)
            print(format_json(data))
            sys.exit(1)
    else:
        excel_bytes = data

    headers, rows, total_rows = _parse_excel_to_rows(
        excel_bytes, max_rows, grep=grep)

    if not rows:
        if grep:
            print(f"(无匹配「{grep}」的数据，共 {total_rows} 条)")
        else:
            print("(无数据)")
        return

    if fmt == "json":
        out = {"total": total_rows, "rows": rows}
        if grep:
            out["grep"] = grep
            out["matched"] = len(rows)
        print(format_json(out))
        return

    shown = len(rows)
    if grep:
        print(f"\n匹配「{grep}」{shown} 条（总 {total_rows} 条）:\n")
    elif total_rows > shown:
        print(f"\n共 {total_rows} 条，展示前 {shown} 条:\n")
    else:
        print(f"\n共 {total_rows} 条:\n")

    formatted_rows = []
    for row in rows:
        formatted_rows.append([_format_value(row.get(h)) for h in headers])
    print(format_markdown_table(headers, formatted_rows))


TD_MAINTENANCE_MSG = (
    "TD 系统近期进行了安全升级，相关 API 接口调用能力受到影响，"
    "目前暂无法从服务端访问，仅支持网页端使用。\n"
    "我们正在与相关团队协同推进恢复方案，具体进展会及时同步。"
)


def _cmd_setup(args):
    staff_id = os.environ.get("TYCLAW_SENDER_STAFF_ID", "")
    if not staff_id:
        print("Error: TYCLAW_SENDER_STAFF_ID env var required", file=sys.stderr)
        sys.exit(1)

    days, expire_time = check_td_token_expiry(args.token)
    if days is not None and days <= 0:
        print(f"Error: Token 已过期（{expire_time}）")
        print("请退出 TD 平台后重新登录，再复制新的 Token。")
        sys.exit(1)

    save_user_credentials(staff_id, "td", {"token": args.token})
    sync_credential_env("td", {"token": args.token})
    print(f"TD credentials saved for {staff_id}")


def _cmd_clear_credentials(_args):
    staff_id = os.environ.get("TYCLAW_SENDER_STAFF_ID", "")
    if not staff_id:
        print("Error: TYCLAW_SENDER_STAFF_ID env var required", file=sys.stderr)
        sys.exit(1)
    if clear_user_credentials(staff_id, "td"):
        clear_credential_env("td")
        print(f"TD credentials cleared for {staff_id}")
    else:
        print(f"No TD credentials found for {staff_id}")


def main():
    # 凭证管理命令不受维护模式影响
    if len(sys.argv) >= 2 and sys.argv[1] in ("setup", "clear-credentials"):
        parser = argparse.ArgumentParser(description="TD credential management")
        sub = parser.add_subparsers(dest="_sub")
        p_setup = sub.add_parser("setup", help="Set TD credentials")
        p_setup.add_argument("--token", required=True)
        sub.add_parser("clear-credentials", help="Clear TD credentials")
        args = parser.parse_args()
        if args._sub == "setup":
            _cmd_setup(args)
        elif args._sub == "clear-credentials":
            _cmd_clear_credentials(args)
        return

    print(TD_MAINTENANCE_MSG, file=sys.stderr)
    sys.exit(1)

    parser = argparse.ArgumentParser(description="TD report query tool")
    parser.add_argument("--config", help="Path to config.yaml")

    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--report", action="store_true",
                      help="Query regular report (platform endpoint)")
    mode.add_argument("--hourly", action="store_true",
                      help="Query hourly report")
    mode.add_argument("--download", action="store_true",
                      help="Download report (channel/channelDetail)")

    parser.add_argument("--body", required=True,
                        help="JSON request body")
    parser.add_argument("--report-type",
                        choices=["channel", "channelDetail"],
                        default="channel",
                        help="Download report type (default: channel)")
    parser.add_argument("--format", choices=["json", "markdown"],
                        default="markdown", dest="fmt",
                        help="Output format (default: markdown)")
    parser.add_argument("--max-length", type=int, default=0,
                        help="Max output characters (0 = no limit)")
    parser.add_argument("--max-rows", type=int, default=50,
                        help="Max rows for download mode (default: 50)")
    parser.add_argument("--grep", default=None,
                        help="Filter rows by keyword (matches any column)")

    args = parser.parse_args()
    staff_id = os.environ.get("TYCLAW_SENDER_STAFF_ID", "")
    config = load_user_config(staff_id, args.config)

    if args.report:
        query_report(config, args.body, args.fmt, args.max_length)
    elif args.hourly:
        query_hourly(config, args.body, args.fmt, args.max_length)
    elif args.download:
        query_download(config, args.body, args.report_type,
                       args.fmt, args.max_rows, grep=args.grep)


if __name__ == "__main__":
    main()
